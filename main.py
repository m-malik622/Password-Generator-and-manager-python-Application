import tkinter as tk
from tkinter import ttk
import string
import random
import openpyxl
window1 = tk.Tk()
window1.title("Password generator")
window1.geometry("800x600")
style = ttk.Style(window1)
style.theme_use('clam')

title_label = ttk.Label(master = window1, text = 'Hello and welcome to the random password \ngenerator. How may I help you')
title_label.pack()

def button1_func():
    def button21_func():
        website = entry21.get()
        length = int(entry22.get())
        window21.destroy()
        window22 = tk.Tk()
        window22.title("Password generator/New file/Finish")
        window22.geometry("800x600")
        if length > 40:
            window22.destroy()
            window23 = tk.Tk()
            window23.title("Password generator/New file/error")
            window23.geometry("800x600")
            title_label221 = ttk.Label(master = window22, text = 'Error: password lenght is too large. Please \nclose tab and try again.')
            title_label221.pack()
            window23.mainloop()

        title_label221 = ttk.Label(master = window22, text = 'The xlsx file has been created and contains \nyour password info. You can now download it and close the window.')
        title_label221.pack()
        
        total = string.ascii_letters + string.digits + string.punctuation
        password = "".join(random.sample(total, length))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Website Name"
        ws.cell(1, 2).value = "Password"
        ws.cell(2, 1).value = website
        ws.cell(2, 2).value = password  
        wb.save("Not_Sus_Doc_At_All.xlsx")
        window22.mainloop()
        return
    

    window1.destroy()
    window21 = tk.Tk()
    window21.title("Password generator/New file")
    window21.geometry("800x600")

    title_label21 = ttk.Label(master = window21, text = 'Please fill in the prompts below')
    title_label21.pack()
    
    prompt_label21 = ttk.Label(master = window21, text = 'Website name')
    prompt_label21.pack()
    entry21 = ttk.Entry(master = window21)
    entry21.pack()
    
    prompt_label22 = ttk.Label(master = window21, text = 'Password length (as a digit)')
    prompt_label22.pack()
    entry22 = ttk.Entry(master = window21)
    entry22.pack()

    button21 = ttk.Button(master = window21, text = 'submit', command = button21_func)
    button21.pack()
    window21.mainloop()
    return


def button2_func():
    def button31_func():
        def button321_func():
            window32.destroy()
            window321 = tk.Tk()
            window321.title("Password generator/Existing file/Overwrite?/Yes")
            window321.geometry("800x600")

            wb = openpyxl.load_workbook("Not_Sus_Doc_At_All.xlsx")
            ws = wb.active 
            last_empty_row = len(list(ws.rows))
            ws.cell(last_empty_row, 1).value = website
            ws.cell(last_empty_row, 2).value = password
            wb.save("Not_Sus_Doc_At_All.xlsx")

            title_label3211 = ttk.Label(master = window321, text = 'The xlsx file has been created and contains \nyour password info. You can now download it and close the window.')
            title_label3211.pack()
            window321.mainloop()
            return
    

        def button322_func():
            def button3221_func():
                file = entry3221.get()
                window322.destroy()
                window3221 = tk.Tk()
                window3221.title("Password generator/Existing file/Overwrite?/No/Finish")
                window3221.geometry("800x600")
                
                wb = openpyxl.load_workbook("Not_Sus_Doc_At_All.xlsx")
                ws = wb.active
                last_empty_row = len(list(ws.rows))
                ws.cell(last_empty_row, 1).value = website
                ws.cell(last_empty_row, 2).value = password
                file_name = file + ".xlsx"
                wb.save(file_name)
                prompt_label3221 = ttk.Label(master = window322, text = 'The xlsx file has been created and contains \nyour password info. You can now download it and close the window.')
                prompt_label3221.pack()
                window3221.mainloop()
                return
            

            window32.destroy()
            window322 = tk.Tk()
            window322.title("Password generator/Existing file/Overwrite?/No")
            window322.geometry("800x600")

            prompt_label3221 = ttk.Label(master = window322, text = 'What would you like the name of this file to be')
            prompt_label3221.pack()
            entry3221 = ttk.Entry(master = window322)
            entry3221.pack()
            button3221 = ttk.Button(master = window322, text = 'Submit', command = button3221_func) 
            button3221.pack()
            window322.mainloop()
            return


        website = entry31.get()
        length = int(entry32.get())
        total = string.ascii_letters + string.digits + string.punctuation
        password = "".join(random.sample(total, length))

        window31.destroy()
        window32 = tk.Tk()
        window32.title("Password generator/Existing file /Overwrite?")
        window32.geometry("800x600")
        if length > 40:
            window32.destroy()
            window33 = tk.Tk()
            window33.title("Password generator/Existing file/error")
            window33.geometry("800x600")
            title_label221 = ttk.Label(master = window33, text = 'Error: password lenght is too large. Please \nclose tab and try again.')
            title_label221.pack()
            window33.mainloop()

        title_label311 = ttk.Label(master = window32, text = 'Would you like to overwrite your current file \n or create a new file')
        title_label311.pack()
        button321 = ttk.Button(master = window32, text = 'Overwrite', command = button321_func) 
        button321.pack()
        button322 = ttk.Button(master = window32, text = 'New file', command = button322_func)
        button322.pack()
        window32.mainloop()
        return
    window1.destroy()
    window31 = tk.Tk()
    window31.title("Password generator/Existing file")
    window31.geometry("800x600")

    title_label31 = ttk.Label(master = window31, text = 'Please fill in the prompts below')
    title_label31.pack()
    prompt_label31 = ttk.Label(master = window31, text = 'Website name')
    prompt_label31.pack()
    entry31 = ttk.Entry(master = window31)
    entry31.pack()
    
    prompt_label32 = ttk.Label(master = window31, text = 'Password length (as a digit)')
    prompt_label32.pack()
    entry32 = ttk.Entry(master = window31)
    entry32.pack()

    button31 = ttk.Button(master = window31, text = 'submit', command = button31_func)
    button31.pack()
    window31.mainloop()    
    return


def button3_func():   
    window1.destroy()
    window41 = tk.Tk()
    window41.title("Password generator/Help")
    window41.geometry("800x600")

    title_label4l = ttk.Label(master = window41, text = "This application is for creating and storing \npasswords.Simply just type the website \nname and password lenght and leave the rest to this \napplication. If you choose to create a new file, the website name and randomly \ngenerated password will appear on a new file generated \nby the code which you could download for later use. \nIf you would like to store the website and password info into an existing \nfile, right-click the random password tab to your left and add \nthe file of choice. You can choose to overwrite your current \nfile or create a new one. If anything ever goes wrong, \ndont feel worried to close the close the \napplication, nothing will happen to your file.")
    title_label4l.pack()
    title_label42 = ttk.Label(master = window41, text = "You may now close this window and restart the program")
    title_label42.pack()
    window41.mainloop()
    return


button11 = ttk.Button(master = window1, text = "I would to create a file with my password info", command = button1_func)
button11.pack()
button12 = ttk.Button(master = window1, text = "I would to add new data to an existing file", command = button2_func)
button12.pack()
button13 = ttk.Button(master = window1, text = "Help", command = button3_func)
button13.pack()
window1.mainloop()